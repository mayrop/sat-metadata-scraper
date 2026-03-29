#!/usr/bin/env python3
"""Download SNICE LIGIE XLSX and build a unified TIGIE hierarchy CSV."""

from __future__ import annotations

import argparse
import csv
import hashlib
import html
import json
import re
import ssl
import sys
import urllib.request
import xml.etree.ElementTree as ET
import zipfile
from collections import defaultdict
from datetime import UTC, datetime
from io import BytesIO, StringIO
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import openpyxl

XLSX_URL = "https://www.snice.gob.mx/~oracle/SNICE_DOCS/TIGIE-MAYO24-TIGIE_20240529-20240529.XLSX"
CHAPTERS_CSV = Path("static/tigie/secciones_capitulos.csv")
HF_DIR = Path("hf/csv/extra/tigie")
HF_EXTRA_DIR = Path("hf/extra/tigie")
DEFAULT_OUTPUT_CSV = HF_DIR / "tigie.csv"
DEFAULT_OUTPUT_XLSX = HF_EXTRA_DIR / Path(XLSX_URL).name
DEFAULT_OUTPUT_JSON = Path("output/tigie-metadata.json")
MANIFEST = Path("output/tigie-manifest.json")
COMBINED_SHEET_CANDIDATES = ("TIGIE + NICO", "TIGIE+NICO")
NICO_ONLY_SHEET_CANDIDATES = ("NICO (ÚNICAMENTE)", "NICO")
CHILD_SHEET_CANDIDATES = ("TICO", "NICO")
USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/145.0.0.0 Safari/537.36"
)


def _sha256_text(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def _sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _write_text_if_changed(path: Path, text: str) -> bool:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists() and path.read_text(encoding="utf-8") == text:
        return False
    path.write_text(text, encoding="utf-8")
    return True


def _write_bytes_if_changed(path: Path, data: bytes) -> bool:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists() and path.read_bytes() == data:
        return False
    path.write_bytes(data)
    return True


def _write_csv_if_changed(path: Path, fieldnames: list[str], rows: list[dict]) -> bool:
    buf = StringIO()
    writer = csv.DictWriter(buf, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)
    return _write_text_if_changed(path, buf.getvalue())


def _write_xlsx(path: Path, fieldnames: list[str], rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "tigie"
    ws.append(fieldnames)
    text_columns = {"A", "B", "C", "Q", "R", "S"}
    for cell in ws[1]:
        cell.number_format = "@"
    for row in rows:
        ws.append([row.get(field, "") for field in fieldnames])
    for col in text_columns:
        for cell in ws[col]:
            cell.number_format = "@"
    wb.save(path)


def _resolve_output_xlsx_path(requested_path: Path, source_file: Path) -> Path:
    if requested_path == DEFAULT_OUTPUT_XLSX:
        return HF_EXTRA_DIR / source_file.name
    return requested_path


def _cleanup_legacy_output_xlsx(resolved_path: Path) -> None:
    if resolved_path != DEFAULT_OUTPUT_XLSX and DEFAULT_OUTPUT_XLSX.exists():
        DEFAULT_OUTPUT_XLSX.unlink()


def _fetch(url: str) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(req, timeout=60, context=ssl._create_unverified_context()) as response:
        return response.read()


def _load_chapter_names(path: Path) -> dict[str, str]:
    with path.open(encoding="utf-8", newline="") as f:
        rows = csv.DictReader(f)
        return {str(int(row["capitulo"])).zfill(2): row["nombre_capitulo"].strip() for row in rows if row.get("capitulo")}


def _normalize_digits(value: str) -> str:
    return re.sub(r"\D", "", value or "")


def _col_idx(ref: str) -> int:
    letters = "".join(ch for ch in ref if ch.isalpha())
    value = 0
    for ch in letters:
        value = value * 26 + (ord(ch.upper()) - ord("A") + 1)
    return value - 1


def _xlsx_shared_strings(zf: zipfile.ZipFile) -> list[str]:
    ns = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    if "xl/sharedStrings.xml" not in zf.namelist():
        return []
    root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
    strings: list[str] = []
    for si in root.findall("a:si", ns):
        strings.append("".join((t.text or "") for t in si.iterfind(".//a:t", ns)))
    return strings


def _xlsx_sheet_rows(data: bytes, sheet_name: str) -> list[list[str]]:
    ns = {
        "a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
    }
    with zipfile.ZipFile(BytesIO(data)) as zf:
        sst = _xlsx_shared_strings(zf)
        workbook = ET.fromstring(zf.read("xl/workbook.xml"))
        rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        rel_by_id = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels}
        target = None
        for sheet in workbook.findall("a:sheets/a:sheet", ns):
            if sheet.attrib["name"] == sheet_name:
                rel_id = sheet.attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"]
                target = rel_by_id[rel_id]
                break
        if not target:
            raise ValueError(f"Sheet {sheet_name!r} not found")
        ws = ET.fromstring(zf.read(f"xl/{target}"))
        rows: list[list[str]] = []
        for row in ws.findall("a:sheetData/a:row", ns):
            values: dict[int, str] = {}
            for cell in row.findall("a:c", ns):
                idx = _col_idx(cell.attrib.get("r", ""))
                value = ""
                v = cell.find("a:v", ns)
                if v is not None and v.text is not None:
                    value = v.text
                    if cell.attrib.get("t") == "s":
                        value = sst[int(value)]
                values[idx] = html.unescape(value).strip()
            width = max(values) + 1 if values else 0
            rows.append([values.get(i, "").strip() for i in range(width)])
        return rows


def _parse_fracciones(rows: list[list[str]], source_url: str) -> list[dict[str, str]]:
    header_idx = next(
        idx
        for idx, row in enumerate(rows)
        if len(row) >= 7
        and _normalize_header(row[2]) == "fracción arancelaria"
        and _normalize_header(row[3]) == "descripción"
    )
    out: list[dict[str, str]] = []
    for row in rows[header_idx + 2 :]:
        if len(row) < 7 or not row[2]:
            continue
        out.append(
            {
                "fraccion_arancelaria": row[2],
                "descripcion": row[3],
                "unidad_medida": row[4],
                "arancel_importacion": row[5],
                "arancel_exportacion": row[6],
                "source_url": source_url,
            }
        )
    if not out:
        raise ValueError("No fracciones parsed")
    return out


def _format_clave(digits: str) -> str:
    if len(digits) == 2:
        return digits
    if len(digits) == 4:
        return digits
    if len(digits) == 6:
        return f"{digits[:4]}.{digits[4:6]}"
    if len(digits) == 8:
        return f"{digits[:4]}.{digits[4:6]}.{digits[6:8]}"
    if len(digits) == 10:
        return f"{digits[:4]}.{digits[4:6]}.{digits[6:8]}.{digits[8:10]}"
    return digits


def _level_info(digits: str) -> tuple[int, str]:
    mapping = {
        2: (1, "capitulo"),
        4: (2, "partida"),
        5: (3, "intermedio"),
        6: (3, "subpartida"),
        8: (4, "fraccion"),
        10: (5, "nico"),
    }
    if len(digits) not in mapping:
        raise ValueError(f"Unsupported TIGIE code length: {digits!r}")
    return mapping[len(digits)]


def _code_sort_key(code: str) -> tuple:
    digits = _normalize_digits(code)
    if not digits:
        return (999, code)
    return tuple(int(digits[i : i + 2]) for i in range(0, len(digits), 2))


def _path_codes(digits: str) -> list[str]:
    parts = [digits[:2], digits[:4], digits[:6], digits[:8], digits[:10]]
    return [_format_clave(part) for part in parts if len(part) <= len(digits) and part]


def _row_hash(row: dict) -> str:
    fields = [
        "clave",
        "parent_clave",
        "nivel",
        "tipo_nivel",
        "nombre",
        "hijos",
        "metadato",
        "source_sheet",
        "source_row_number",
        "fraccion_arancelaria",
        "nico",
        "unidad_medida",
        "arancel_importacion",
        "arancel_exportacion",
        "fa_correlativa",
        "nico_correlativa",
    ]
    return _sha256_text("|".join(str(row.get(field, "")) for field in fields))


def _normalize_header(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").replace("\n", " ").strip()).lower()


def _parse_child_rows(rows: list[list[str]], source_url: str, code_key: str) -> list[dict[str, str]]:
    header_idx = next(
        idx
        for idx, row in enumerate(rows)
        if len(row) >= 6
        and _normalize_header(row[3]) == "fracción arancelaria"
        and code_key in _normalize_header(row[4])
        and _normalize_header(row[5]) == "descripción"
    )
    out: list[dict[str, str]] = []
    for row in rows[header_idx + 1 :]:
        if len(row) < 6 or not row[3]:
            continue
        item = {
            "fraccion_arancelaria": row[3],
            "nico": row[4],
            "descripcion": row[5],
            "source_url": source_url,
        }
        out.append(item)
    if not out:
        raise ValueError(f"No {code_key.upper()} rows parsed")
    return out


def _load_first_sheet(data: bytes, candidates: tuple[str, ...]) -> tuple[str, list[list[str]]]:
    for name in candidates:
        try:
            return name, _xlsx_sheet_rows(data, name)
        except ValueError:
            continue
    raise ValueError(f"None of the expected child sheets were found: {', '.join(candidates)}")


def _find_parent_from_stack(stack: list[tuple[str, str]], digits: str) -> str:
    while stack:
        parent_clave, parent_digits = stack[-1]
        if len(parent_digits) < len(digits) and digits.startswith(parent_digits):
            return parent_clave
        stack.pop()
    return ""


def _build_paths(rows_by_code: dict[str, dict[str, str]]) -> None:
    def path_for(clave: str) -> list[str]:
        out: list[str] = []
        seen: set[str] = set()
        current = clave
        while current and current not in seen and current in rows_by_code:
            seen.add(current)
            out.append(current)
            current = rows_by_code[current]["parent_clave"]
        out.reverse()
        return out

    for clave, row in rows_by_code.items():
        row["path_claves_json"] = json.dumps(path_for(clave), ensure_ascii=False)


def _parse_combined_rows(rows: list[list[str]], source_url: str) -> list[dict[str, str]]:
    header_idx = next(
        idx
        for idx, row in enumerate(rows)
        if len(row) >= 10
        and _normalize_header(row[1]) == "fracción arancelaria"
        and _normalize_header(row[2]) == "nico"
        and _normalize_header(row[3]) == "descripción"
    )
    out: list[dict[str, str]] = []
    for row_num, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        fraccion = row[1].strip() if len(row) > 1 else ""
        nico = row[2].strip() if len(row) > 2 else ""
        descripcion = row[3].strip() if len(row) > 3 else ""
        if not (fraccion or nico or descripcion):
            continue
        if not descripcion:
            continue
        out.append(
            {
                "fraccion_arancelaria": fraccion,
                "nico": nico,
                "nico_columna": nico,
                "descripcion": descripcion,
                "fa_correlativa": row[4].strip() if len(row) > 4 else "",
                "nico_correlativa": row[5].strip() if len(row) > 5 else "",
                "unidad_medida": row[6].strip() if len(row) > 6 else "",
                "arancel_importacion": row[7].strip() if len(row) > 7 else "",
                "arancel_exportacion": row[8].strip() if len(row) > 8 else "",
                "observaciones_implementacion_septima_enmienda": row[9].strip() if len(row) > 9 else "",
                "modificaciones_respecto_a_ligie_enviada": row[10].strip() if len(row) > 10 else "",
                "source_url": source_url,
                "__source_row_number": str(row_num),
            }
        )
    if not out:
        raise ValueError("No combined TIGIE rows parsed")
    return out


def _parse_nico_only_rows(rows: list[list[str]], source_url: str, sheet_name: str) -> list[dict[str, str]]:
    if sheet_name == "NICO (ÚNICAMENTE)":
        header_idx = next(
            idx
            for idx, row in enumerate(rows)
            if len(row) >= 4
            and _normalize_header(row[1]) == "fracción arancelaria"
            and _normalize_header(row[2]) == "nico"
            and _normalize_header(row[3]) == "descripción"
        )
        start_col = 1
    else:
        header_idx = next(
            idx
            for idx, row in enumerate(rows)
            if len(row) >= 6
            and _normalize_header(row[3]) == "fracción arancelaria"
            and _normalize_header(row[4]) == "nico"
            and _normalize_header(row[5]) == "descripción"
        )
        start_col = 3
    out: list[dict[str, str]] = []
    for row_num, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        fraccion = row[start_col].strip() if len(row) > start_col else ""
        nico = row[start_col + 1].strip() if len(row) > start_col + 1 else ""
        descripcion = row[start_col + 2].strip() if len(row) > start_col + 2 else ""
        if not (fraccion or nico or descripcion):
            continue
        if not nico.isdigit():
            continue
        out.append(
            {
                "fraccion_arancelaria": fraccion,
                "nico": nico,
                "nico_columna": nico,
                "descripcion": descripcion,
                "source_url": source_url,
                "__source_row_number": str(row_num),
            }
        )
    if not out:
        raise ValueError(f"No NICO rows parsed from {sheet_name}")
    return out


def _build_parent_row(
    digits: str,
    source_file: Path,
    source_url: str,
    raw_hash: str,
    path_codes: list[str],
    nombre: str = "",
) -> dict[str, str]:
    nivel, tipo_nivel = _level_info(digits)
    clave = _format_clave(digits)
    parent_digits = digits[:-2]
    return {
        "clave": clave,
        "clave_normalizada": digits,
        "parent_clave": _format_clave(parent_digits) if parent_digits else "",
        "nivel": str(nivel),
        "tipo_nivel": tipo_nivel,
        "nombre": nombre,
        "nombre_ing": "",
        "hijos": "0",
        "metadato": "1",
        "es_hoja": "0",
        "tabindex": "",
        "source_file": str(source_file),
        "source_sheet": "DERIVED",
        "source_row_number": "",
        "source_url": source_url,
        "raw_hash": raw_hash,
        "path_claves_json": json.dumps(path_codes, ensure_ascii=False),
        "fraccion_arancelaria": "",
        "nico": "",
        "nico_columna": "",
        "descripcion": "",
        "fa_correlativa": "",
        "nico_correlativa": "",
        "unidad_medida": "",
        "arancel_importacion": "",
        "arancel_exportacion": "",
        "observaciones_implementacion_septima_enmienda": "",
        "modificaciones_respecto_a_ligie_enviada": "",
    }


def _build_fraccion_row(
    row: dict[str, str],
    source_file: Path,
    source_url: str,
    raw_hash: str,
    source_row_number: int,
) -> dict[str, str]:
    digits = _normalize_digits(row["fraccion_arancelaria"])
    nivel, tipo_nivel = _level_info(digits)
    return {
        "clave": _format_clave(digits),
        "clave_normalizada": digits,
        "parent_clave": _format_clave(digits[:-2]),
        "nivel": str(nivel),
        "tipo_nivel": tipo_nivel,
        "nombre": row["descripcion"],
        "nombre_ing": "",
        "hijos": "0",
        "metadato": "0",
        "es_hoja": "0",
        "tabindex": "",
        "source_file": str(source_file),
        "source_sheet": "FA",
        "source_row_number": str(source_row_number),
        "source_url": source_url,
        "raw_hash": raw_hash,
        "path_claves_json": json.dumps(_path_codes(digits), ensure_ascii=False),
        "fraccion_arancelaria": row["fraccion_arancelaria"],
        "nico": "",
        "nico_columna": row.get("nico_columna", ""),
        "descripcion": row["descripcion"],
        "fa_correlativa": row.get("fa_correlativa", row["fraccion_arancelaria"]),
        "nico_correlativa": "",
        "unidad_medida": row["unidad_medida"],
        "arancel_importacion": row["arancel_importacion"],
        "arancel_exportacion": row["arancel_exportacion"],
        "observaciones_implementacion_septima_enmienda": row.get("observaciones_implementacion_septima_enmienda", ""),
        "modificaciones_respecto_a_ligie_enviada": row.get("modificaciones_respecto_a_ligie_enviada", ""),
    }


def _build_nico_row(
    row: dict[str, str],
    source_file: Path,
    source_url: str,
    raw_hash: str,
    source_row_number: int,
    child_label: str,
) -> dict[str, str]:
    fraccion_digits = _normalize_digits(row["fraccion_arancelaria"])
    digits = f"{fraccion_digits}{_normalize_digits(row['nico']).zfill(2)}"
    nivel, _ = _level_info(digits)
    return {
        "clave": _format_clave(digits),
        "clave_normalizada": digits,
        "parent_clave": _format_clave(fraccion_digits),
        "nivel": str(nivel),
        "tipo_nivel": child_label,
        "nombre": row["descripcion"],
        "nombre_ing": "",
        "hijos": "0",
        "metadato": "0",
        "es_hoja": "1",
        "tabindex": "",
        "source_file": str(source_file),
        "source_sheet": child_label.upper(),
        "source_row_number": str(source_row_number),
        "source_url": source_url,
        "raw_hash": raw_hash,
        "path_claves_json": json.dumps(_path_codes(digits), ensure_ascii=False),
        "fraccion_arancelaria": row["fraccion_arancelaria"],
        "nico": row["nico"],
        "nico_columna": row.get("nico_columna", row["nico"]),
        "descripcion": row["descripcion"],
        "fa_correlativa": row.get("fa_correlativa", row["fraccion_arancelaria"]),
        "nico_correlativa": row.get("nico_correlativa", ""),
        "unidad_medida": "",
        "arancel_importacion": "",
        "arancel_exportacion": "",
        "observaciones_implementacion_septima_enmienda": row.get("observaciones_implementacion_septima_enmienda", ""),
        "modificaciones_respecto_a_ligie_enviada": row.get("modificaciones_respecto_a_ligie_enviada", ""),
    }


def _enrich_rows(rows_by_code: dict[str, dict[str, str]]) -> list[dict[str, str]]:
    child_counts: dict[str, int] = defaultdict(int)
    for row in rows_by_code.values():
        parent = row["parent_clave"]
        if parent:
            child_counts[parent] += 1

    for code, row in rows_by_code.items():
        row["hijos"] = "1" if child_counts.get(code, 0) else "0"
        row["es_hoja"] = "0" if child_counts.get(code, 0) else "1"
        row["row_hash"] = _row_hash(row)
    return sorted(rows_by_code.values(), key=lambda item: (_code_sort_key(item["clave"]), item["clave"]))


def _build_unified_rows_from_combined(
    combined_rows: list[dict[str, str]],
    source_file: Path,
    source_url: str,
    raw_hash: str,
    chapter_names: dict[str, str],
) -> list[dict[str, str]]:
    rows_by_code: dict[str, dict[str, str]] = {}
    stack: list[tuple[str, str]] = []
    current_fraccion_clave = ""

    for item in combined_rows:
        fraccion = item["fraccion_arancelaria"]
        nico = item["nico"]
        numeric_nico = nico.isdigit()
        if fraccion:
            fr_digits = _normalize_digits(fraccion)
            fr_clave = fraccion

            if len(fr_digits) >= 2:
                chapter_clave = _format_clave(fr_digits[:2])
                rows_by_code.setdefault(
                    chapter_clave,
                    _build_parent_row(
                        fr_digits[:2],
                        source_file,
                        source_url,
                        raw_hash,
                        [chapter_clave],
                        chapter_names.get(fr_digits[:2], ""),
                    ),
                )

            parent_clave = _find_parent_from_stack(stack, fr_digits)
            if not parent_clave and len(fr_digits) >= 4:
                parent_clave = _format_clave(fr_digits[:2])

            nivel, tipo_nivel = _level_info(fr_digits)
            rows_by_code[fr_clave] = {
                "clave": fr_clave,
                "clave_normalizada": fr_digits,
                "parent_clave": parent_clave,
                "nivel": str(nivel),
                "tipo_nivel": tipo_nivel,
                "nombre": item["descripcion"],
                "nombre_ing": "",
                "hijos": "0",
                "metadato": "0" if len(fr_digits) >= 4 else "1",
                "es_hoja": "0",
                "tabindex": "",
                "source_file": str(source_file),
                "source_sheet": "TIGIE + NICO",
                "source_row_number": item["__source_row_number"],
                "source_url": source_url,
                "raw_hash": raw_hash,
                "path_claves_json": "[]",
                "fraccion_arancelaria": fraccion,
                "nico": "",
                "nico_columna": item["nico_columna"],
                "descripcion": item["descripcion"],
                "fa_correlativa": item["fa_correlativa"],
                "nico_correlativa": "",
                "unidad_medida": item["unidad_medida"],
                "arancel_importacion": item["arancel_importacion"],
                "arancel_exportacion": item["arancel_exportacion"],
                "observaciones_implementacion_septima_enmienda": item["observaciones_implementacion_septima_enmienda"],
                "modificaciones_respecto_a_ligie_enviada": item["modificaciones_respecto_a_ligie_enviada"],
            }
            while stack and not (len(stack[-1][1]) < len(fr_digits) and fr_digits.startswith(stack[-1][1])):
                stack.pop()
            stack.append((fr_clave, fr_digits))
            current_fraccion_clave = fr_clave if len(fr_digits) >= 8 else current_fraccion_clave

        if numeric_nico:
            nico_clave = item["nico_correlativa"] or f"{(fraccion or current_fraccion_clave)}.{nico}"
            parent_clave = fraccion or current_fraccion_clave
            if not parent_clave:
                raise ValueError(f"NICO row without active fraccion at source row {item['__source_row_number']}")
            fraccion_digits = _normalize_digits(parent_clave)
            nico_row = _build_nico_row(
                {
                    "fraccion_arancelaria": parent_clave,
                    "nico": nico,
                    "nico_columna": item["nico_columna"],
                    "descripcion": item["descripcion"],
                    "fa_correlativa": item["fa_correlativa"],
                    "nico_correlativa": nico_clave,
                    "observaciones_implementacion_septima_enmienda": item["observaciones_implementacion_septima_enmienda"],
                    "modificaciones_respecto_a_ligie_enviada": item["modificaciones_respecto_a_ligie_enviada"],
                },
                source_file,
                source_url,
                raw_hash,
                int(item["__source_row_number"]),
                "nico",
            )
            nico_row["clave"] = nico_clave
            nico_row["clave_normalizada"] = _normalize_digits(nico_clave)
            nico_row["parent_clave"] = parent_clave
            rows_by_code[nico_clave] = nico_row
            current_fraccion_clave = fraccion or current_fraccion_clave

    _build_paths(rows_by_code)
    return _enrich_rows(rows_by_code)


def _extract_nico_metadata_from_combined(combined_rows: list[dict[str, str]]) -> dict[str, dict[str, str]]:
    out: dict[str, dict[str, str]] = {}
    current_fraccion_clave = ""
    for item in combined_rows:
        fraccion = item["fraccion_arancelaria"]
        nico = item["nico"]
        if fraccion:
            fr_digits = _normalize_digits(fraccion)
            if len(fr_digits) >= 8:
                current_fraccion_clave = fraccion
        if not nico.isdigit():
            continue
        parent_clave = fraccion or current_fraccion_clave
        if not parent_clave:
            continue
        clave = item["nico_correlativa"] or f"{parent_clave}.{nico}"
        out[clave] = {
            "fa_correlativa": item.get("fa_correlativa", ""),
            "nico_correlativa": item.get("nico_correlativa", clave),
            "unidad_medida": item.get("unidad_medida", ""),
            "arancel_importacion": item.get("arancel_importacion", ""),
            "arancel_exportacion": item.get("arancel_exportacion", ""),
            "observaciones_implementacion_septima_enmienda": item.get("observaciones_implementacion_septima_enmienda", ""),
            "modificaciones_respecto_a_ligie_enviada": item.get("modificaciones_respecto_a_ligie_enviada", ""),
        }
    return out


def _replace_nicos_from_sheet(
    unified_rows: list[dict[str, str]],
    nico_rows: list[dict[str, str]],
    source_file: Path,
    source_url: str,
    raw_hash: str,
    metadata_by_clave: dict[str, dict[str, str]],
    source_sheet_name: str,
) -> list[dict[str, str]]:
    rows_by_code = {row["clave"]: dict(row) for row in unified_rows if row["tipo_nivel"] != "nico"}
    for item in nico_rows:
        clave = f"{item['fraccion_arancelaria']}.{item['nico']}"
        extra = metadata_by_clave.get(clave, {})
        nico_row = _build_nico_row(
            {
                "fraccion_arancelaria": item["fraccion_arancelaria"],
                "nico": item["nico"],
                "nico_columna": item["nico_columna"],
                "descripcion": item["descripcion"],
                "fa_correlativa": extra.get("fa_correlativa", item["fraccion_arancelaria"]),
                "nico_correlativa": extra.get("nico_correlativa", clave),
                "observaciones_implementacion_septima_enmienda": extra.get("observaciones_implementacion_septima_enmienda", ""),
                "modificaciones_respecto_a_ligie_enviada": extra.get("modificaciones_respecto_a_ligie_enviada", ""),
            },
            source_file,
            source_url,
            raw_hash,
            int(item["__source_row_number"]),
            "nico",
        )
        nico_row["clave"] = clave
        nico_row["clave_normalizada"] = _normalize_digits(clave)
        nico_row["parent_clave"] = item["fraccion_arancelaria"]
        nico_row["source_sheet"] = source_sheet_name
        nico_row["source_row_number"] = item["__source_row_number"]
        nico_row["unidad_medida"] = extra.get("unidad_medida", "")
        nico_row["arancel_importacion"] = extra.get("arancel_importacion", "")
        nico_row["arancel_exportacion"] = extra.get("arancel_exportacion", "")
        rows_by_code[clave] = nico_row
    _build_paths(rows_by_code)
    return _enrich_rows(rows_by_code)


def _build_unified_rows(
    fracciones: list[dict[str, str]],
    nicos: list[dict[str, str]],
    source_file: Path,
    source_url: str,
    raw_hash: str,
    child_label: str,
    chapter_names: dict[str, str],
) -> list[dict[str, str]]:
    rows_by_code: dict[str, dict[str, str]] = {}

    for item in fracciones:
        digits = _normalize_digits(item["fraccion_arancelaria"])
        path_codes = _path_codes(digits)
        for ancestor_digits in (digits[:2], digits[:4], digits[:6]):
            rows_by_code.setdefault(
                _format_clave(ancestor_digits),
                _build_parent_row(
                    ancestor_digits,
                    source_file,
                    source_url,
                    raw_hash,
                    _path_codes(ancestor_digits),
                    chapter_names.get(ancestor_digits, "") if len(ancestor_digits) == 2 else "",
                ),
            )
        source_row_number = int(item.get("__source_row_number", "0") or 0)
        rows_by_code[_format_clave(digits)] = _build_fraccion_row(item, source_file, source_url, raw_hash, source_row_number)

    for item in nicos:
        fraccion_digits = _normalize_digits(item["fraccion_arancelaria"])
        for ancestor_digits in (fraccion_digits[:2], fraccion_digits[:4], fraccion_digits[:6]):
            rows_by_code.setdefault(
                _format_clave(ancestor_digits),
                _build_parent_row(
                    ancestor_digits,
                    source_file,
                    source_url,
                    raw_hash,
                    _path_codes(ancestor_digits),
                    chapter_names.get(ancestor_digits, "") if len(ancestor_digits) == 2 else "",
                ),
            )
        source_row_number = int(item.get("__source_row_number", "0") or 0)
        nico_row = _build_nico_row(item, source_file, source_url, raw_hash, source_row_number, child_label)
        rows_by_code[nico_row["clave"]] = nico_row

    return _enrich_rows(rows_by_code)


def _annotate_source_rows(rows: list[dict[str, str]], source_sheet_rows: list[list[str]], sheet_name: str) -> list[dict[str, str]]:
    out: list[dict[str, str]] = []
    data_start = 8 if sheet_name == "FA" else 7
    parsed_idx = 0
    for row_num, raw_row in enumerate(source_sheet_rows, start=1):
        if row_num <= data_start:
            continue
        if sheet_name == "FA":
            if len(raw_row) < 7 or not raw_row[2]:
                continue
        else:
            if len(raw_row) < 6 or not raw_row[3]:
                continue
        if parsed_idx >= len(rows):
            break
        item = dict(rows[parsed_idx])
        item["__source_row_number"] = str(row_num)
        out.append(item)
        parsed_idx += 1
    if parsed_idx != len(rows):
        raise ValueError(f"Annotated {parsed_idx} {sheet_name} rows but parsed {len(rows)}")
    return out


def _load_xlsx_bytes(input_xlsx: Path | None, verify: bool) -> tuple[bytes, Path, str]:
    if input_xlsx is not None:
        data = input_xlsx.read_bytes()
        return data, input_xlsx, XLSX_URL
    data = _fetch(XLSX_URL)
    return data, Path(Path(XLSX_URL).name), XLSX_URL


def build_metadata(rows: list[dict[str, str]], source_file: Path, source_url: str, raw_hash: str) -> dict[str, object]:
    by_type = defaultdict(int)
    for row in rows:
        by_type[row["tipo_nivel"]] += 1
    return {
        "dataset": "tigie",
        "source_url": source_url,
        "source_file": str(source_file),
        "raw_hash": raw_hash,
        "rows": len(rows),
        "tipos": dict(sorted(by_type.items())),
        "leaf_rows": sum(1 for row in rows if row["es_hoja"] == "1"),
        "max_level": max((int(row["nivel"]) for row in rows), default=0),
        "hash": _sha256_text("".join(row["row_hash"] for row in rows)),
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--verify", action="store_true", help="Re-download the SNICE workbook even if cached locally.")
    parser.add_argument("--input-xlsx", type=Path, help="Use a local XLSX file instead of downloading the default SNICE source.")
    parser.add_argument("--output-csv", type=Path, default=DEFAULT_OUTPUT_CSV)
    parser.add_argument("--output-xlsx", type=Path, default=DEFAULT_OUTPUT_XLSX)
    parser.add_argument("--output-json", type=Path, default=DEFAULT_OUTPUT_JSON)
    args = parser.parse_args()

    xlsx_bytes, source_file, source_url = _load_xlsx_bytes(args.input_xlsx, args.verify)
    output_xlsx_path = _resolve_output_xlsx_path(args.output_xlsx, source_file)
    _cleanup_legacy_output_xlsx(output_xlsx_path)
    raw_hash = _sha256_bytes(xlsx_bytes)
    chapter_names = _load_chapter_names(CHAPTERS_CSV)

    combined_rows: list[dict[str, str]] | None = None
    fracciones: list[dict[str, str]] = []
    nicos: list[dict[str, str]] = []
    child_label = "nico"
    try:
        _, combined_sheet_rows = _load_first_sheet(xlsx_bytes, COMBINED_SHEET_CANDIDATES)
        combined_rows = _parse_combined_rows(combined_sheet_rows, source_url)
        unified_rows = _build_unified_rows_from_combined(combined_rows, source_file, source_url, raw_hash, chapter_names)
        nico_sheet_name, nico_only_sheet_rows = _load_first_sheet(xlsx_bytes, NICO_ONLY_SHEET_CANDIDATES)
        nicos = _parse_nico_only_rows(nico_only_sheet_rows, source_url, nico_sheet_name)
        unified_rows = _replace_nicos_from_sheet(
            unified_rows,
            nicos,
            source_file,
            source_url,
            raw_hash,
            _extract_nico_metadata_from_combined(combined_rows),
            nico_sheet_name,
        )
        child_label = "nico"
    except ValueError as exc:
        if "None of the expected child sheets were found" not in str(exc):
            raise
        fa_rows = _xlsx_sheet_rows(xlsx_bytes, "FA")
        child_sheet_name, child_rows = _load_first_sheet(xlsx_bytes, CHILD_SHEET_CANDIDATES)
        child_label = child_sheet_name.lower()
        fracciones = _annotate_source_rows(_parse_fracciones(fa_rows, source_url), fa_rows, "FA")
        nicos = _annotate_source_rows(_parse_child_rows(child_rows, source_url, child_label), child_rows, child_sheet_name)
        unified_rows = _build_unified_rows(fracciones, nicos, source_file, source_url, raw_hash, child_label, chapter_names)
    metadata = build_metadata(unified_rows, source_file, source_url, raw_hash)

    fieldnames = [
        "clave",
        "clave_normalizada",
        "parent_clave",
        "nivel",
        "tipo_nivel",
        "nombre",
        "nombre_ing",
        "hijos",
        "metadato",
        "es_hoja",
        "tabindex",
        "source_file",
        "source_sheet",
        "source_row_number",
        "source_url",
        "raw_hash",
        "path_claves_json",
        "fraccion_arancelaria",
        "nico",
        "nico_columna",
        "descripcion",
        "fa_correlativa",
        "nico_correlativa",
        "unidad_medida",
        "arancel_importacion",
        "arancel_exportacion",
        "observaciones_implementacion_septima_enmienda",
        "modificaciones_respecto_a_ligie_enviada",
        "row_hash",
    ]

    _write_csv_if_changed(args.output_csv, fieldnames, unified_rows)
    _write_xlsx(output_xlsx_path, fieldnames, unified_rows)
    _write_text_if_changed(args.output_json, json.dumps(metadata, ensure_ascii=False, indent=2))
    manifest = {
        "fecha_extraccion": datetime.now(UTC).isoformat(),
        "source_url": source_url,
        "source_file": str(source_file),
        "raw_hash": raw_hash,
        "counts": {
            "fracciones": len(fracciones),
            child_label: len(nicos),
            "combined_rows": len(combined_rows or []),
            "unified_rows": len(unified_rows),
        },
        "files": {
            "raw_xlsx": str(source_file),
            "raw_consolidated": str(args.output_csv),
            "raw_consolidated_xlsx": str(output_xlsx_path),
            "raw_metadata": str(args.output_json),
        },
        "tipos": metadata["tipos"],
    }
    _write_text_if_changed(MANIFEST, json.dumps(manifest, ensure_ascii=False, indent=2))
    print(
        f"Wrote {len(unified_rows)} unified TIGIE rows from {len(combined_rows or fracciones)} source rows",
        file=sys.stderr,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
