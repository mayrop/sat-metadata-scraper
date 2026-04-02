#!/usr/bin/env python3
"""Extract SAT CFDI catalog and matrix XLS files into CSV files.

Reads output/catalog.csv to find the latest XLS catalog files scraped by
scrape.py (those stored under hf/raw/catalogos/).

- `file_type=catalogo.xls` entries are parsed as SAT catalog workbooks, extracting
  every sheet whose name starts with `c_`.
- `file_type=matriz` entries are parsed with matrix-specific header detection.

Example outputs:
    hf/raw/catalogos/anexo20/cfdi/catCFDI40.xls   → hf/csv/anexo20/cfdi/c_uso_cfdi.csv …
    hf/raw/catalogos/complementos/carta-porte/x.xls → hf/csv/complementos/carta-porte/c_estaciones.csv …

Usage:
    uv run scripts/catalogos/extract.py
    uv run scripts/catalogos/extract.py --catalog output/catalog.csv --header-style snake
    uv run scripts/catalogos/extract.py --xls-dir hf/raw/catalogos --csv-dir hf/csv
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import re
import sys
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Dict, Iterable, List, Sequence

import openpyxl
import xlrd

# ── constants ──────────────────────────────────────────────────────────────────

HF_XLS_DIR    = Path("hf/raw/catalogos")
HF_CSV_DIR    = Path("hf/csv")
CATALOG_STATE = Path("catalog_state.csv")  # committed to git, outside hf/
CATALOG_CSV   = Path("output/catalog.csv")

CATALOG_HEADER_PATTERN = re.compile(r"^[cC]_\w+")
PART_SUFFIX_PATTERN    = re.compile(r"(?:_Parte)?_\d+$", re.IGNORECASE)


# ── data types ─────────────────────────────────────────────────────────────────


@dataclass
class SheetExtraction:
    headers: List[str]
    rows: List[List[str]]
    metadata: Dict[str, str]
    description: str


# ── string helpers ─────────────────────────────────────────────────────────────


def normalize_token(value: str) -> str:
    value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    value = value.replace("\xa0", " ")
    value = re.sub(r"\s+", " ", value)
    return value.strip().lower()


SPECIAL_SNAKE_HEADERS = {
    "fechainiciovigencia": "fecha_de_inicio_de_vigencia",
    "fechafinvigencia": "fecha_de_fin_de_vigencia",
}


def slugify(value: str) -> str:
    token = normalize_token(value)
    token = re.sub(r"[^0-9a-z]+", "_", token)
    return token.strip("_")


def strip_catalog_prefix(value: str) -> str:
    return re.sub(r"^(?:c|C)_", "", value.strip())


def split_compound_words(token: str) -> str:
    step1 = re.sub(r"([a-z])([A-Z])", r"\1 \2", token)
    step2 = re.sub(r"([A-Za-z])([0-9])", r"\1 \2", step1)
    step3 = re.sub(r"([0-9])([A-Za-z])", r"\1 \2", step2)
    return step3


def alnum_token(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^0-9a-z]+", "", normalized.lower())


def to_snake_ascii(value: str) -> str:
    normalized_key = alnum_token(value)
    if normalized_key in SPECIAL_SNAKE_HEADERS:
        return SPECIAL_SNAKE_HEADERS[normalized_key]
    ascii_value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    spaced = split_compound_words(ascii_value)
    ascii_value = re.sub(r"[^0-9A-Za-z]+", "_", spaced)
    ascii_value = re.sub(r"_+", "_", ascii_value)
    return ascii_value.strip("_").lower()


def first_non_empty(strings: Sequence[str]) -> str:
    for item in strings:
        if item:
            return item
    return ""


def override_folder_version(section: str, folder_version: str, source_xls: str) -> str:
    source_norm = source_xls.replace("\\", "/").lower()
    if section == "complementos/recepcion-de-pagos" and (
        source_norm.endswith("/cat_pagos.xls")
        or source_norm.endswith("/matriz_de_errores_crp_v20_rev_b.xls")
    ):
        return "2-0"
    return folder_version


def apply_row_overrides(section: str, row: Dict[str, str]) -> Dict[str, str]:
    source_norm = row.get("source_xls", "").replace("\\", "/").lower()
    if section == "complementos/recepcion-de-pagos" and (
        source_norm.endswith("/cat_pagos.xls")
        or source_norm.endswith("/matriz_de_errores_crp_v20_rev_b.xls")
    ):
        updated = dict(row)
        updated["version"] = "2"
        updated["revision"] = "0"
        return updated
    return row


def normalize_matrix_catalog_name(stem: str) -> str:
    stem = re.sub(r"_(?:\d{8}|[0-9a-f]{8,})$", "", stem, flags=re.IGNORECASE)
    return stem


def normalize_section(category: str, slug: str) -> str:
    if category in {"complementos_retenciones", "complementos-retenciones"}:
        return "/".join(part for part in ["complementos-retenciones", slug] if part)
    if category == "anexo-20":
        if slug == "formato-de-factura":
            slug = "factura-electronica"
        elif slug == "factura-de-retenciones-e-informacion-de-pagos":
            slug = "factura-de-retenciones"
    if category in {"complementos-concepto", "anexo-20"}:
        return "/".join(part for part in [category, slug] if part)
    return "/".join(part for part in [slugify(category), slug] if part)


def normalize_folder_version(folder_version: str) -> str:
    value = folder_version or "files"
    if value.startswith("version-"):
        value = value[len("version-") :]
    value = value.replace("-revision-", "-")
    if value.startswith("revision-"):
        value = value[len("revision-") :]
    return value


def folder_version_from_local_file(local_file: str) -> str:
    for part in Path(local_file).parts:
        if part.startswith("version-"):
            return normalize_folder_version(part[len("version-") :])
    return "files"


# ── XLS parsing ────────────────────────────────────────────────────────────────


def sheet_is_empty(row_values: Iterable[str]) -> bool:
    return not any(cell for cell in row_values)


def is_continuation_row(row_values: Sequence[str]) -> bool:
    first = first_non_empty(row_values)
    if not first:
        return False
    return normalize_token(first).startswith("continua en")


def format_number(value: float) -> str:
    if float(value).is_integer():
        return str(int(value))
    return ("%s" % value).rstrip("0").rstrip(".")


def is_xlsx_workbook(workbook: Any) -> bool:
    return isinstance(workbook, openpyxl.Workbook)


def sheet_nrows(sheet: Any) -> int:
    return sheet.max_row if hasattr(sheet, "max_row") else sheet.nrows


def sheet_ncols(sheet: Any) -> int:
    return sheet.max_column if hasattr(sheet, "max_column") else sheet.ncols


def sheet_row_len(sheet: Any, row_idx: int) -> int:
    if hasattr(sheet, "max_column"):
        return sheet.max_column
    return sheet.row_len(row_idx)


def raw_cell_value(sheet: Any, row_idx: int, col_idx: int) -> Any:
    if hasattr(sheet, "cell") and hasattr(sheet, "max_row"):
        return sheet.cell(row_idx + 1, col_idx + 1).value
    return sheet.cell(row_idx, col_idx).value


def row_values_raw(sheet: Any, row_idx: int) -> list[Any]:
    width = sheet_row_len(sheet, row_idx)
    return [raw_cell_value(sheet, row_idx, col_idx) for col_idx in range(width)]


def sheet_label(sheet: Any) -> str:
    return getattr(sheet, "name", getattr(sheet, "title", "<unknown>"))


def raw_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def format_cell(workbook: Any, sheet: Any, row_idx: int, col_idx: int) -> str:
    if is_xlsx_workbook(workbook):
        cell = sheet.cell(row_idx + 1, col_idx + 1)
        value = cell.value
        if value is None:
            return ""
        if isinstance(value, datetime):
            if value.time() == datetime.min.time():
                return value.date().isoformat()
            return value.isoformat()
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, time):
            return value.isoformat()
        if isinstance(value, bool):
            return "TRUE" if value else "FALSE"
        if isinstance(value, (int, float)):
            fmt_str = (cell.number_format or "").split(";")[0].strip()
            if fmt_str and re.fullmatch(r"0+", fmt_str):
                width = len(fmt_str)
                return f"{int(round(float(value))):0{width}d}"
            return format_number(float(value))
        if cell.data_type == "e":
            return "#ERROR"
        return str(value).strip()

    cell = sheet.cell(row_idx, col_idx)
    if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
        return ""
    if cell.ctype == xlrd.XL_CELL_DATE:
        dt = xlrd.xldate_as_datetime(cell.value, workbook.datemode)
        if dt.time() == datetime.min.time():
            return dt.date().isoformat()
        return dt.isoformat()
    if cell.ctype == xlrd.XL_CELL_NUMBER:
        xf = workbook.xf_list[cell.xf_index]
        fmt = workbook.format_map.get(xf.format_key)
        fmt_str = fmt.format_str.split(";")[0] if fmt else ""
        fmt_str = fmt_str.strip()
        if fmt_str and re.fullmatch(r"0+", fmt_str):
            width = len(fmt_str)
            return f"{int(round(cell.value)):0{width}d}"
        return format_number(cell.value)
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return "TRUE" if cell.value else "FALSE"
    if cell.ctype == xlrd.XL_CELL_ERROR:
        return "#ERROR"
    return str(cell.value).strip()


def detect_catalog_header_row(sheet: xlrd.sheet.Sheet) -> int:
    """Find the header row index.

    Primary strategy: look for a row where any cell starts with 'c_' (Anexo 20 style).
    Fallback: find the last empty-row separator in the first 10 rows, then return
    the next non-empty row after it (Carta Porte / complement-style sheets).
    """
    for row_idx in range(sheet_nrows(sheet)):
        row_cells = [raw_text(value) for value in row_values_raw(sheet, row_idx)]
        if sheet_is_empty(row_cells):
            continue
        if any(CATALOG_HEADER_PATTERN.match(value) for value in row_cells if value):
            return row_idx
        normalized = [normalize_token(value) for value in row_cells if value]
        if (
            len(normalized) >= 2
            and any("descripcion" in value for value in normalized)
            and not any(value == "version" or "fecha de publicacion" in value for value in normalized)
        ):
            return row_idx

    # Fallback: header follows the last empty separator in the metadata block
    last_empty = None
    for row_idx in range(min(10, sheet_nrows(sheet))):
        row_cells = [raw_text(value) for value in row_values_raw(sheet, row_idx)]
        if sheet_is_empty(row_cells):
            last_empty = row_idx
    if last_empty is not None:
        for row_idx in range(last_empty + 1, sheet_nrows(sheet)):
            row_cells = [raw_text(value) for value in row_values_raw(sheet, row_idx)]
            if not sheet_is_empty(row_cells):
                return row_idx

    raise ValueError(f"Header row not found in sheet {sheet_label(sheet)!r}")


def gather_header_rows(sheet: xlrd.sheet.Sheet, start_idx: int) -> List[int]:
    rows = [start_idx]
    row_idx = start_idx + 1
    while row_idx < sheet_nrows(sheet):
        row_values = [raw_text(value) for value in row_values_raw(sheet, row_idx)]
        if sheet_is_empty(row_values):
            row_idx += 1
            continue
        first_cell = raw_text(raw_cell_value(sheet, row_idx, 0)) if sheet_ncols(sheet) else ""
        if not first_cell:
            rows.append(row_idx)
            row_idx += 1
            continue
        break
    return rows


def combine_headers(
    workbook: Any, sheet: Any, header_rows: List[int]
) -> List[tuple[int, str]]:
    width = max(sheet_row_len(sheet, idx) for idx in header_rows)
    headers: List[tuple[int, str]] = []
    for col_idx in range(width):
        parts: List[str] = []
        for row_idx in header_rows:
            if col_idx >= sheet_row_len(sheet, row_idx):
                continue
            value = format_cell(workbook, sheet, row_idx, col_idx)
            if value:
                parts.append(value)
        if not parts:
            continue
        header_value = " - ".join(parts).strip() or f"column_{col_idx + 1}"
        headers.append((col_idx, header_value))
    return headers


def find_description(sheet: xlrd.sheet.Sheet, workbook: xlrd.book.Book) -> str:
    for row_idx in range(min(6, sheet_nrows(sheet))):
        values = [
            format_cell(workbook, sheet, row_idx, col_idx)
            for col_idx in range(sheet_row_len(sheet, row_idx))
        ]
        cleaned = [val for val in values if val]
        if cleaned and any("catalogo" in normalize_token(val) for val in cleaned):
            return " | ".join(cleaned)
    return ""


_METADATA_KEY_ALIASES: Dict[str, str] = {
    # version
    "version_catalogo":                       "version",
    "version_cfdi":                           "version",
    # revision
    "revision_catalogo":                      "revision",
    # fecha publicacion
    "fecha_publicacion_de_catalogo":          "fecha_publicacion",
    "fecha_de_publicacion":                   "fecha_publicacion",
    # fecha inicio vigencia
    "fecha_inicio_de_vigencia_del_catalogo":  "fecha_inicio_vigencia",
    "fecha_inicio_de_vigencia":               "fecha_inicio_vigencia",
    "fecha_de_inicio_de_vigencia":            "fecha_inicio_vigencia",
    "fechainiciodevigencia":                  "fecha_inicio_vigencia",
    "fechainiciovigencia":                    "fecha_inicio_vigencia",
    # fecha fin vigencia
    "fecha_fin_de_vigencia_del_catalogo":     "fecha_fin_vigencia",
    "fecha_fin_de_vigencia":                  "fecha_fin_vigencia",
    "fecha_de_fin_de_vigencia":               "fecha_fin_vigencia",
    "fechafindevigencia":                     "fecha_fin_vigencia",
    "fechafinvigencia":                       "fecha_fin_vigencia",
}


def extract_metadata(sheet: xlrd.sheet.Sheet, workbook: xlrd.book.Book) -> Dict[str, str]:
    key_row_idx = None
    for row_idx in range(min(10, sheet_nrows(sheet))):
        for col_idx in range(sheet_row_len(sheet, row_idx)):
            if "version" in normalize_token(raw_text(raw_cell_value(sheet, row_idx, col_idx))):
                key_row_idx = row_idx
                break
        if key_row_idx is not None:
            break
    if key_row_idx is None:
        return {}
    value_row_idx = key_row_idx + 1
    while value_row_idx < sheet_nrows(sheet) and sheet_row_len(sheet, value_row_idx) == 0:
        value_row_idx += 1
    if value_row_idx >= sheet_nrows(sheet):
        return {}
    metadata: Dict[str, str] = {}
    for col_idx in range(sheet_row_len(sheet, key_row_idx)):
        key_raw = raw_text(raw_cell_value(sheet, key_row_idx, col_idx))
        if not key_raw:
            continue
        key = slugify(key_raw)
        if not key:
            continue
        key = _METADATA_KEY_ALIASES.get(key, key)
        value = format_cell(workbook, sheet, value_row_idx, col_idx)
        if "fecha" in key and value == "0":
            value = ""
        metadata[key] = value
    return metadata


def extract_data_rows(
    workbook: Any,
    sheet: Any,
    header_rows: List[int],
    header_columns: List[tuple[int, str]],
) -> List[List[str]]:
    first_data_row = max(header_rows) + 1
    column_indices = [col_idx for col_idx, _ in header_columns]
    rows: List[List[str]] = []
    row_idx = first_data_row
    while row_idx < sheet_nrows(sheet):
        row_values = [format_cell(workbook, sheet, row_idx, col_idx) for col_idx in column_indices]
        if sheet_is_empty(row_values):
            row_idx += 1
            continue
        if is_continuation_row(row_values):
            row_idx += 1
            continue
        rows.append(row_values)
        row_idx += 1
    return _merge_orphan_rows(rows)


def _merge_orphan_rows(rows: List[List[str]]) -> List[List[str]]:
    """Merge rows with an empty first column into the previous row.

    When the XLS has a multi-value cell (e.g. N/Nómina with valor_maximo=NS on
    one row and the numeric value on the next), the continuation row has an empty
    clave. We merge it back by replacing any cell in the previous row that is
    empty or equals 'NS' with the non-empty value from the continuation row.
    """
    merged: List[List[str]] = []
    for row in rows:
        if row and not row[0] and merged:
            prev = merged[-1]
            for i, val in enumerate(row):
                if i < len(prev) and val and (not prev[i] or prev[i] == "NS"):
                    prev[i] = val
        else:
            merged.append(row)
    return merged


def parse_sheet(workbook: Any, sheet: Any) -> SheetExtraction:
    header_start = detect_catalog_header_row(sheet)
    header_rows = gather_header_rows(sheet, header_start)
    header_columns = combine_headers(workbook, sheet, header_rows)
    headers = [header for _, header in header_columns]
    rows = extract_data_rows(workbook, sheet, header_rows, header_columns)
    metadata = extract_metadata(sheet, workbook)
    description = find_description(sheet, workbook)
    return SheetExtraction(headers=headers, rows=rows, metadata=metadata, description=description)


def base_catalog_name(sheet_name: str) -> str:
    name = PART_SUFFIX_PATTERN.sub("", sheet_name)
    if name.startswith("C_"):
        name = "c_" + name[2:]
    return name


def transform_headers_for_output(headers: Sequence[str], catalog: str, style: str) -> List[str]:
    if style == "original":
        return list(headers)
    if style == "snake":
        catalog_snake = to_snake_ascii(catalog)
        catalog_snake_stripped = to_snake_ascii(strip_catalog_prefix(catalog))
        transformed: List[str] = []
        for idx, header in enumerate(headers, start=1):
            snake = to_snake_ascii(strip_catalog_prefix(header)) or f"column_{idx}"
            if snake in {catalog_snake, catalog_snake_stripped}:
                snake = "clave"
            transformed.append(snake)
        return transformed
    raise ValueError(f"Unsupported header style: {style}")


def write_csv(path: Path, headers: Sequence[str], rows: Sequence[Sequence[str]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        writer.writerows(rows)


# ── per-workbook extraction ────────────────────────────────────────────────────


def extract_workbook(xls_path: Path, output_dir: Path, header_style: str) -> tuple[List[str], List[Dict]]:
    """Extract all c_* sheets from xls_path into output_dir.

    Returns (list of written csv paths, list of metadata dicts).
    Metadata is not written here — caller accumulates across workbooks and writes once per dir.
    """
    if xls_path.suffix.lower() == ".xlsx":
        workbook = openpyxl.load_workbook(xls_path, data_only=True)
        sheet_names = workbook.sheetnames
        get_sheet = workbook.__getitem__
    else:
        workbook = xlrd.open_workbook(str(xls_path), formatting_info=True)
        sheet_names = workbook.sheet_names()
        get_sheet = workbook.sheet_by_name
    catalogs: Dict[str, Dict] = defaultdict(
        lambda: {"headers": None, "rows": [], "metadata": {}, "description": "", "sheets": []}
    )
    for sheet_name in sheet_names:
        if not sheet_name.lower().startswith("c_"):
            continue
        sheet = get_sheet(sheet_name)
        try:
            parsed = parse_sheet(workbook, sheet)
        except ValueError as exc:
            print(f"  SKIP {sheet_name!r}: {exc}", file=sys.stderr)
            continue
        # For single-catalog XLS files (stem starts with c_), use the XLS stem
        # as the catalog name so that e.g. c_FraccionArancelaria_v17_rA.xls
        # produces c_FraccionArancelaria_v17_rA.csv instead of colliding with
        # c_FraccionArancelaria.csv.
        xls_stem = xls_path.stem
        if xls_stem.lower().startswith("c_"):
            base_name = xls_stem
        else:
            base_name = base_catalog_name(sheet_name)
        record = catalogs[base_name]
        if record["headers"] is None:
            record["headers"] = parsed.headers
        elif record["headers"] != parsed.headers:
            print(f"  WARN header mismatch for {base_name!r} — skipping extra sheet", file=sys.stderr)
            continue
        record["rows"].extend(parsed.rows)
        if not record["metadata"] and parsed.metadata:
            record["metadata"] = parsed.metadata
        if not record["description"] and parsed.description:
            record["description"] = parsed.description
        record["sheets"].append(sheet_name)

    output_dir.mkdir(parents=True, exist_ok=True)
    written: List[str] = []
    metadata_rows: List[Dict[str, str]] = []

    for catalog, payload in catalogs.items():
        if payload["headers"] is None:
            continue
        out_headers = transform_headers_for_output(payload["headers"], catalog, header_style) + ["row_hash"]
        dest = output_dir / f"{catalog}.csv"
        hashed_rows = [
            row + [hashlib.sha256("|".join(row).encode()).hexdigest()]
            for row in payload["rows"]
        ]
        write_csv(dest, out_headers, hashed_rows)
        print(f"  → {dest}  ({len(payload['rows'])} rows)", file=sys.stderr)
        written.append(str(dest))

        entry: Dict[str, str] = {
            "catalogo": catalog,
            "source_xls": str(xls_path),
            "xls_hash": hashlib.sha256(xls_path.read_bytes()).hexdigest(),
            "descripcion": payload["description"],
            "sheets": ",".join(payload["sheets"]),
            "numero_columnas": str(len(out_headers)),
            "nombres_columnas": "|".join(out_headers),
            "file_hash": hashlib.sha256(dest.read_bytes()).hexdigest(),
            "file_type": "catalogo",
        }
        entry.update(payload["metadata"])
        metadata_rows.append(entry)

    return written, metadata_rows


def load_matrix_workbook(path: Path) -> tuple[Any, Any]:
    if path.suffix.lower() == ".xlsx":
        workbook = openpyxl.load_workbook(path, data_only=True)
        return workbook, workbook[workbook.sheetnames[0]]
    workbook = xlrd.open_workbook(str(path), formatting_info=True)
    return workbook, workbook.sheet_by_index(0)


def detect_matrix_header_row(workbook: Any, sheet: Any) -> int:
    for row_idx in range(min(15, sheet_nrows(sheet))):
        values = [format_cell(workbook, sheet, row_idx, col_idx) for col_idx in range(sheet_ncols(sheet))]
        normalized = [normalize_token(v) for v in values if v]
        joined = " ".join(normalized)
        has_codigo_error = "codigo error" in joined or "codigo de error" in joined
        if has_codigo_error and ("validacion" in joined or "error" in joined):
            return row_idx
        if (
            has_codigo_error
            and "regla de validacion" in joined
            and ("descripcion del error" in joined or "aclaraciones" in joined)
        ):
            return row_idx
    raise ValueError("Could not detect matrix header row")


def detect_matrix_used_columns(workbook: Any, sheet: Any, header_row: int) -> list[int]:
    used: list[int] = []
    for col_idx in range(sheet_ncols(sheet)):
        has_value = False
        for row_idx in range(header_row, sheet_nrows(sheet)):
            if format_cell(workbook, sheet, row_idx, col_idx):
                has_value = True
                break
        if has_value:
            used.append(col_idx)
    return used


def build_matrix_headers(workbook: Any, sheet: Any, header_row: int, used_columns: list[int]) -> list[str]:
    headers: list[str] = []
    extra_counter = 1
    last_used = used_columns[-1] if used_columns else -1
    for col_idx in used_columns:
        value = format_cell(workbook, sheet, header_row, col_idx)
        if value:
            headers.append(slugify(value) or f"column_{col_idx + 1}")
            continue
        if col_idx == last_used:
            headers.append("notas")
        else:
            headers.append(f"extra_{extra_counter}")
            extra_counter += 1
    return headers


def extract_matrix_rows(workbook: Any, sheet: Any, header_row: int, used_columns: list[int]) -> list[list[str]]:
    rows: list[list[str]] = []
    for row_idx in range(header_row + 1, sheet_nrows(sheet)):
        row = [format_cell(workbook, sheet, row_idx, col_idx) for col_idx in used_columns]
        if not any(row):
            continue
        rows.append(row)
    return rows


def extract_matrix_file(source_path: Path, output_dir: Path) -> dict[str, str]:
    workbook, sheet = load_matrix_workbook(source_path)
    header_row = detect_matrix_header_row(workbook, sheet)
    used_columns = detect_matrix_used_columns(workbook, sheet, header_row)
    headers = build_matrix_headers(workbook, sheet, header_row, used_columns)
    rows = extract_matrix_rows(workbook, sheet, header_row, used_columns)
    out_rows = [r + [hashlib.sha256("|".join(r).encode()).hexdigest()] for r in rows]
    out_headers = headers + ["row_hash"]
    catalogo = normalize_matrix_catalog_name(source_path.stem)
    dest = output_dir / f"{catalogo}.csv"
    output_dir.mkdir(parents=True, exist_ok=True)
    write_csv(dest, out_headers, out_rows)
    print(f"  → {dest}  ({len(rows)} rows)", file=sys.stderr)
    return {
        "catalogo": catalogo,
        "source_xls": str(source_path),
        "xls_hash": hashlib.sha256(source_path.read_bytes()).hexdigest(),
        "descripcion": normalize_token(
            format_cell(workbook, sheet, 0, 0) if sheet_nrows(sheet) else ""
        ),
        "sheets": getattr(sheet, "name", getattr(sheet, "title", "")),
        "numero_columnas": str(len(out_headers)),
        "nombres_columnas": "|".join(out_headers),
        "file_hash": hashlib.sha256(dest.read_bytes()).hexdigest(),
        "file_type": "matriz",
    }


# ── discovery ──────────────────────────────────────────────────────────────────


def discover_xls(xls_dir: Path, catalog_file: Path) -> List[Path]:
    """Return catalog XLS/XLSX paths referenced in output/catalog.csv."""
    files: List[Path] = []
    seen: set[Path] = set()
    if not catalog_file.exists():
        return files
    with catalog_file.open(newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            if row.get("file_type") != "catalogo.xls":
                continue
            source_xls = row.get("local_file", "")
            if not source_xls:
                continue
            path = Path(source_xls)
            if not path.exists():
                continue
            if xls_dir not in path.parents and path != xls_dir:
                continue
            if path in seen:
                continue
            seen.add(path)
            files.append(path)
    return sorted(files)


def _logical_rel_parent(
    xls_path: Path,
    xls_dir: Path,
    source_section_overrides: dict[str, str],
) -> Path:
    def _normalize_parent_name(path: Path) -> Path:
        parts = list(path.parts)
        if parts and parts[-1].startswith("version-"):
            parts[-1] = normalize_folder_version(parts[-1][len("version-"):])
        return Path(*parts) if parts else path

    try:
        override_rel = source_section_overrides.get(str(xls_path))
        if override_rel:
            parent_name = xls_path.parent.name
            if parent_name.startswith("version-"):
                parent_name = normalize_folder_version(parent_name[len("version-"):])
            return Path(override_rel) / parent_name
        return _normalize_parent_name(xls_path.parent.relative_to(xls_dir))
    except ValueError:
        return _normalize_parent_name(xls_path.parent)


def discover_matrix_rows(catalog_file: Path) -> list[dict[str, str]]:
    if not catalog_file.exists():
        return []
    with catalog_file.open(newline="", encoding="utf-8") as f:
        return [dict(row) for row in csv.DictReader(f) if row.get("file_type") == "matriz"]


# ── CLI ────────────────────────────────────────────────────────────────────────


def parse_args(argv: Sequence[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--xls-dir",
        type=Path,
        default=HF_XLS_DIR,
        help=f"Root directory of downloaded XLS files (default: {HF_XLS_DIR})",
    )
    parser.add_argument(
        "--csv-dir",
        type=Path,
        default=HF_CSV_DIR,
        help=f"Root output directory for extracted CSVs (default: {HF_CSV_DIR})",
    )
    parser.add_argument(
        "--state-file",
        type=Path,
        default=CATALOG_STATE,
        help=f"Path to catalog state CSV (default: {CATALOG_STATE})",
    )
    parser.add_argument(
        "--header-style",
        choices=["original", "snake"],
        default="snake",
        help="Column name format: 'original' keeps SAT names, 'snake' converts to snake_case (default: snake)",
    )
    parser.add_argument(
        "--catalog-file",
        type=Path,
        default=CATALOG_CSV,
        help=f"Path to scraped catalog CSV for section overrides (default: {CATALOG_CSV})",
    )
    parser.add_argument(
        "--sections", nargs="+", metavar="SECTION",
        help=(
            "Only extract XLS files under these section paths. "
            "Examples: anexo20/cfdi  anexo20/retenciones  complementos"
        ),
    )
    return parser.parse_args(argv)


def _state_key(row: dict[str, str]) -> tuple[str, str, str]:
    return (
        row.get("section", ""),
        normalize_folder_version(row.get("folder_version", "")),
        row.get("catalogo", ""),
    )


def _load_catalog_state(state_file: Path) -> dict[tuple[str, str, str], dict]:
    """Load catalog_state.csv into a dict keyed by logical catalog identity."""
    state: dict[tuple[str, str, str], dict] = {}
    if not state_file.exists():
        return state
    with state_file.open(newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            row = dict(row)
            row["folder_version"] = normalize_folder_version(row.get("folder_version", ""))
            key = _state_key(row)
            state[key] = row
    return state


def _prefer_state_row(current: dict[str, str], candidate: dict[str, str]) -> dict[str, str]:
    def score(row: dict[str, str]) -> tuple[int, int, int]:
        folder_version = row.get("folder_version", "")
        return (
            1 if folder_version and folder_version != "files" else 0,
            1 if row.get("version") else 0,
            1 if row.get("revision") else 0,
        )

    return candidate if score(candidate) >= score(current) else current


def _dedupe_merged_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    deduped: dict[tuple[str, str, str, str], dict[str, str]] = {}
    for row in rows:
        key = (
            row.get("section", ""),
            row.get("catalogo", ""),
            row.get("source_xls", ""),
            row.get("file_type", ""),
        )
        if key in deduped:
            deduped[key] = _prefer_state_row(deduped[key], row)
        else:
            deduped[key] = row
    return sorted(
        deduped.values(),
        key=lambda r: (r.get("section", ""), r.get("folder_version", ""), r.get("catalogo", "")),
    )


def _load_source_section_overrides(catalog_file: Path) -> dict[str, str]:
    """Map source XLS paths to desired CSV section paths using output/catalog.csv."""
    overrides: dict[str, str] = {}
    if not catalog_file.exists():
        return overrides
    with catalog_file.open(newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            if row.get("file_type") != "catalogo.xls":
                continue
            source_xls = row.get("local_file", "")
            category = row.get("category", "")
            slug = row.get("slug", "")
            if not source_xls or not category:
                continue
            if category == "complementos-concepto" and slug:
                overrides[source_xls] = f"complementos-concepto/{slug}"
    return overrides


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv or sys.argv[1:])

    xls_files = discover_xls(args.xls_dir, args.catalog_file)
    matrix_rows = discover_matrix_rows(args.catalog_file)
    source_section_overrides = _load_source_section_overrides(args.catalog_file)
    if args.sections:
        sections_filter = [s.strip("/") for s in args.sections]
        xls_files = [
            p for p in xls_files
            if any(
                str(_logical_rel_parent(p, args.xls_dir, source_section_overrides)).startswith(s + "/") or
                str(_logical_rel_parent(p, args.xls_dir, source_section_overrides)).startswith(s + "\\") or
                str(_logical_rel_parent(p, args.xls_dir, source_section_overrides)) == s
                for s in sections_filter
            )
        ]
        matrix_rows = [
            row for row in matrix_rows
            if any(
                normalize_section(row.get("category", ""), row.get("slug", "")).startswith(s)
                for s in sections_filter
            )
        ]
        print(f"Filtering to sections: {sections_filter}", file=sys.stderr)
    if not xls_files and not matrix_rows:
        print("No XLS files found — skipping extract.", file=sys.stderr)
        return 0

    # Load existing catalog_state.csv for merge
    existing_state = _load_catalog_state(args.state_file)

    print(f"Found {len(xls_files)} catalog XLS file(s) and {len(matrix_rows)} matrix file(s):", file=sys.stderr)
    total_written = 0
    failed_files: List[tuple[Path, str]] = []
    # Accumulate new metadata rows keyed by (source_xls, catalogo)
    new_state: dict[tuple[str, str, str], dict] = {}
    # Accumulate per output dir so multiple XLS in the same folder merge correctly
    metadata_by_dir: dict[Path, List[Dict]] = defaultdict(list)

    for xls_path in xls_files:
        if not xls_path.exists():
            print(f"\n[{xls_path}] not downloaded (version unchanged) — skipping", file=sys.stderr)
            continue
        print(f"\n[{xls_path}]", file=sys.stderr)
        rel_parent = _logical_rel_parent(xls_path, args.xls_dir, source_section_overrides)
        rel_parts = rel_parent.parts
        rel_folder_version = rel_parts[-1] if rel_parts else ""
        rel_section = "/".join(rel_parts[:-1]) if len(rel_parts) > 1 else str(rel_parent)
        effective_rel_folder_version = override_folder_version(
            rel_section,
            rel_folder_version,
            str(xls_path),
        )
        if effective_rel_folder_version != rel_folder_version:
            rel_parent = Path(rel_section) / effective_rel_folder_version
        output_dir = args.csv_dir / rel_parent
        try:
            written, meta_rows = extract_workbook(xls_path, output_dir, args.header_style)
        except Exception as exc:
            failed_files.append((xls_path, str(exc)))
            print(f"  ERROR {xls_path}: {exc}", file=sys.stderr)
            continue
        total_written += len(written)
        metadata_by_dir[output_dir].extend(meta_rows)

    for row in matrix_rows:
        source_rel = row.get("local_file", "")
        if not source_rel:
            continue
        source_path = Path(source_rel)
        if not source_path.exists():
            failed_files.append((source_path, "Missing source matrix"))
            print(f"  ERROR {source_path}: Missing source matrix", file=sys.stderr)
            continue
        section = normalize_section(row.get("category", ""), row.get("slug", ""))
        folder_version = folder_version_from_local_file(source_rel)
        output_dir = args.csv_dir / section / folder_version
        try:
            entry = extract_matrix_file(source_path, output_dir)
        except Exception as exc:
            failed_files.append((source_path, str(exc)))
            print(f"  ERROR {source_path}: {exc}", file=sys.stderr)
            continue
        entry["version"] = row.get("version", "")
        entry["revision"] = row.get("revision", "")
        entry["fecha_publicacion"] = row.get("last_modified", "")
        metadata_by_dir[output_dir].append(entry)
        total_written += 1

    # Build new index rows from freshly processed files, adding section + version from path
    for output_dir, meta_rows in metadata_by_dir.items():
        try:
            rel = output_dir.relative_to(args.csv_dir)
        except ValueError:
            rel = output_dir
        parts = rel.parts
        folder_version = parts[-1] if parts else ""
        section = "/".join(parts[:-1]) if len(parts) > 1 else str(rel)
        for row in meta_rows:
            row = apply_row_overrides(section, row)
            effective_folder_version = override_folder_version(
                section,
                folder_version,
                row.get("source_xls", ""),
            )
            entry = {"section": section, "folder_version": effective_folder_version, **row}
            key = _state_key(entry)
            new_state[key] = entry

    # Merge: start from existing state, overwrite with any freshly processed entries
    merged: dict[tuple, dict] = {**existing_state, **new_state}

    # Determine column order (union of all keys, stable)
    index_keys: List[str] = ["section", "folder_version"]
    for row in merged.values():
        for k in row:
            if k not in index_keys:
                index_keys.append(k)

    merged_rows = _dedupe_merged_rows(list(merged.values()))
    if merged_rows:
        args.state_file.parent.mkdir(parents=True, exist_ok=True)
        write_csv(args.state_file, index_keys, [[r.get(k, "") for k in index_keys] for r in merged_rows])
        updated = len(new_state)
        preserved = len(merged_rows) - updated
        print(
            f"\nState   → {args.state_file}  ({len(merged_rows)} entries, {updated} updated, {preserved} preserved)",
            file=sys.stderr,
        )

    if failed_files:
        print("\nErrors:", file=sys.stderr)
        for path, message in failed_files:
            print(f"  - {path}: {message}", file=sys.stderr)

    print(f"Done. Wrote {total_written} CSV file(s) to {args.csv_dir}/", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
