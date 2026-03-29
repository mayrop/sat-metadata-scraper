#!/usr/bin/env python3
"""Extract SAT matrix XLS/XLSX files into per-matrix CSV files.

Reads output/catalog.csv rows with file_type=matriz and writes CSV outputs under
hf/csv/, plus a separate matrix_state.csv that can be merged into the main HF dataset.

Usage:
    uv run scripts/catalogos/extract_matrices.py
    uv run scripts/catalogos/extract_matrices.py --csv-dir hf/csv --state-file matrix_state.csv
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import re
import sys
import unicodedata
from pathlib import Path
from typing import Any, Sequence

import openpyxl
import xlrd

CATALOG_CSV = Path("output/catalog.csv")
CSV_DIR = Path("hf/csv")
STATE_FILE = Path("matrix_state.csv")


def normalize_token(value: str) -> str:
    value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    value = value.replace("\xa0", " ")
    value = re.sub(r"\s+", " ", value)
    return value.strip().lower()


def slugify(value: str) -> str:
    token = normalize_token(value)
    token = re.sub(r"[^0-9a-z]+", "_", token)
    return token.strip("_")


def normalize_matrix_catalog_name(stem: str) -> str:
    stem = re.sub(r"_(?:\d{8}|[0-9a-f]{8,})$", "", stem, flags=re.IGNORECASE)
    return stem


def format_number(value: float) -> str:
    if float(value).is_integer():
        return str(int(value))
    return ("%s" % value).rstrip("0").rstrip(".")


def is_xlsx_workbook(workbook: Any) -> bool:
    return isinstance(workbook, openpyxl.Workbook)


def sheet_nrows(sheet: Any) -> int:
    return sheet.max_row if hasattr(sheet, "max_row") else sheet.nrows


def sheet_ncols(sheet: Any) -> int:
    return sheet.max_column if hasattr(sheet, "max_column") else sheet.ncols


def raw_cell_value(sheet: Any, row_idx: int, col_idx: int) -> Any:
    if hasattr(sheet, "cell") and hasattr(sheet, "max_row"):
        return sheet.cell(row_idx + 1, col_idx + 1).value
    return sheet.cell(row_idx, col_idx).value


def format_cell(workbook: Any, sheet: Any, row_idx: int, col_idx: int) -> str:
    if is_xlsx_workbook(workbook):
        cell = sheet.cell(row_idx + 1, col_idx + 1)
        value = cell.value
        if value is None:
            return ""
        if isinstance(value, bool):
            return "TRUE" if value else "FALSE"
        if isinstance(value, (int, float)):
            return format_number(float(value))
        return str(value).strip()

    cell = sheet.cell(row_idx, col_idx)
    if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
        return ""
    if cell.ctype == xlrd.XL_CELL_NUMBER:
        return format_number(cell.value)
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return "TRUE" if cell.value else "FALSE"
    if cell.ctype == xlrd.XL_CELL_ERROR:
        return "#ERROR"
    return str(cell.value).strip()


def load_workbook(path: Path) -> tuple[Any, Any]:
    if path.suffix.lower() == ".xlsx":
        workbook = openpyxl.load_workbook(path, data_only=True)
        return workbook, workbook[workbook.sheetnames[0]]
    workbook = xlrd.open_workbook(str(path), formatting_info=False)
    return workbook, workbook.sheet_by_index(0)


def detect_header_row(workbook: Any, sheet: Any) -> int:
    for row_idx in range(min(15, sheet_nrows(sheet))):
        values = [format_cell(workbook, sheet, row_idx, col_idx) for col_idx in range(sheet_ncols(sheet))]
        normalized = [normalize_token(v) for v in values if v]
        joined = " ".join(normalized)
        if "codigo error" in joined and ("validacion" in joined or "error" in joined):
            return row_idx
    raise ValueError("Could not detect matrix header row")


def detect_used_columns(workbook: Any, sheet: Any, header_row: int) -> list[int]:
    used: list[int] = []
    for col_idx in range(sheet_ncols(sheet)):
        has_value = False
        for row_idx in range(header_row, sheet_nrows(sheet)):
            if format_cell(workbook, sheet, row_idx, col_idx):
                has_value = True
                break
        if has_value:
            used.append(col_idx)
    return used


def build_headers(workbook: Any, sheet: Any, header_row: int, used_columns: list[int]) -> list[str]:
    headers: list[str] = []
    extra_counter = 1
    last_used = used_columns[-1] if used_columns else -1
    for col_idx in used_columns:
        value = format_cell(workbook, sheet, header_row, col_idx)
        if value:
            headers.append(slugify(value) or f"column_{col_idx + 1}")
            continue
        if col_idx == last_used:
            headers.append("notas")
        else:
            headers.append(f"extra_{extra_counter}")
            extra_counter += 1
    return headers


def extract_rows(workbook: Any, sheet: Any, header_row: int, used_columns: list[int]) -> list[list[str]]:
    rows: list[list[str]] = []
    for row_idx in range(header_row + 1, sheet_nrows(sheet)):
        row = [format_cell(workbook, sheet, row_idx, col_idx) for col_idx in used_columns]
        if not any(row):
            continue
        rows.append(row)
    return rows


def normalize_section(category: str, slug: str) -> str:
    if category == "anexo-20":
        if slug == "formato-de-factura":
            return "anexo20/cfdi"
        if slug == "factura-de-retenciones-e-informacion-de-pagos":
            return "anexo20/retenciones"
    return "/".join(part for part in [slugify(category), slug] if part)


def folder_version_from_local_file(local_file: str) -> str:
    for part in Path(local_file).parts:
        if part.startswith("version-"):
            return part[len("version-") :]
    return "files"


def discover_matrix_rows(catalog_csv: Path) -> list[dict[str, str]]:
    with catalog_csv.open(newline="", encoding="utf-8") as handle:
        return [dict(row) for row in csv.DictReader(handle) if row.get("file_type") == "matriz"]


def write_csv(path: Path, headers: Sequence[str], rows: Sequence[Sequence[str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        writer.writerows(rows)


def load_state(path: Path) -> dict[str, dict[str, str]]:
    if not path.exists():
        return {}
    with path.open(newline="", encoding="utf-8") as handle:
        return {
            row.get("source_xls", ""): dict(row)
            for row in csv.DictReader(handle)
            if row.get("source_xls", "")
        }


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--catalog-file", type=Path, default=CATALOG_CSV)
    parser.add_argument("--csv-dir", type=Path, default=CSV_DIR)
    parser.add_argument("--state-file", type=Path, default=STATE_FILE)
    args = parser.parse_args(argv)

    matrix_rows = discover_matrix_rows(args.catalog_file)
    if not matrix_rows:
        print("No matrix rows found in output/catalog.csv", file=sys.stderr)
        return 0

    existing_state = load_state(args.state_file)
    new_state: dict[str, dict[str, str]] = {}
    written = 0

    for row in matrix_rows:
        source_rel = row.get("local_file", "")
        if not source_rel:
            continue
        source_path = Path("output/files") / source_rel
        if not source_path.exists():
            print(f"Missing source matrix: {source_path}", file=sys.stderr)
            continue

        xls_hash = hashlib.sha256(source_path.read_bytes()).hexdigest()
        catalogo = normalize_matrix_catalog_name(Path(source_path).stem)
        key = str(source_path)
        section = normalize_section(row.get("category", ""), row.get("slug", ""))
        folder_version = folder_version_from_local_file(source_rel)
        dest = args.csv_dir / section / folder_version / f"{catalogo}.csv"

        workbook, sheet = load_workbook(source_path)
        header_row = detect_header_row(workbook, sheet)
        used_columns = detect_used_columns(workbook, sheet, header_row)
        headers = build_headers(workbook, sheet, header_row, used_columns)
        rows = extract_rows(workbook, sheet, header_row, used_columns)
        out_rows = [r + [hashlib.sha256("|".join(r).encode()).hexdigest()] for r in rows]
        out_headers = headers + ["row_hash"]

        write_csv(dest, out_headers, out_rows)
        written += 1
        print(f"  → {dest}  ({len(rows)} rows)", file=sys.stderr)

        new_state[key] = {
            "section": section,
            "folder_version": folder_version,
            "catalogo": catalogo,
            "source_xls": str(source_path),
            "xls_hash": xls_hash,
            "descripcion": normalize_token(
                format_cell(workbook, sheet, 0, 0) if sheet_nrows(sheet) else ""
            ),
            "sheets": getattr(sheet, "name", getattr(sheet, "title", "")),
            "numero_columnas": str(len(out_headers)),
            "nombres_columnas": "|".join(out_headers),
            "file_hash": hashlib.sha256(dest.read_bytes()).hexdigest(),
            "version": row.get("version", ""),
            "revision": row.get("revision", ""),
            "fecha_publicacion": row.get("last_modified", ""),
            "file_type": "matriz",
        }

    merged = {**existing_state, **new_state}
    fieldnames = [
        "section",
        "folder_version",
        "catalogo",
        "source_xls",
        "xls_hash",
        "descripcion",
        "sheets",
        "numero_columnas",
        "nombres_columnas",
        "file_hash",
        "version",
        "revision",
        "fecha_publicacion",
        "file_type",
    ]
    rows_out = [
        [row.get(field, "") for field in fieldnames]
        for row in sorted(
            merged.values(),
            key=lambda r: (r.get("section", ""), r.get("folder_version", ""), r.get("catalogo", "")),
        )
    ]
    write_csv(args.state_file, fieldnames, rows_out)
    print(f"State   → {args.state_file}  ({len(rows_out)} matrices)", file=sys.stderr)
    print(f"Done. Wrote {written} matrix CSV file(s) to {args.csv_dir}/", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
